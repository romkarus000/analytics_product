from __future__ import annotations

import csv
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import CurrentUser
from app.db.session import get_db
from app.models.column_mapping import ColumnMapping
from app.models.fact_marketing_spend import FactMarketingSpend
from app.models.fact_transaction import FactTransaction
from app.models.project import Project
from app.models.project_settings import ProjectSettings
from app.models.upload import Upload, UploadStatus, UploadType
from app.models.upload_quarantine import UploadQuarantineRow
from app.schemas.uploads import (
    ColumnMappingCreate,
    ColumnMappingPublic,
    ImportResult,
    QualityIssue,
    QualityReport,
    QualityStats,
    UploadPreview,
)
from app.services.upload_pipeline import (
    build_field_mapping,
    extract_operation_type_mapping,
    extract_unknown_operation_policy,
    get_row_value,
    normalize_value,
    parse_date,
    parse_float,
    read_upload_rows,
)
from app.services.insights import generate_insights_for_project
from app.services.aliases import (
    get_manager_name,
    get_product_name,
    resolve_manager_alias,
    resolve_product_alias,
)

router = APIRouter(prefix="/uploads", tags=["upload-mapping"])

REQUIRED_FIELDS: dict[UploadType, list[str]] = {
    UploadType.TRANSACTIONS: [
        "paid_at",
        "operation_type",
        "amount",
    ],
    UploadType.MARKETING_SPEND: ["date", "spend_amount"],
}

SUGGESTION_RULES: dict[str, list[str]] = {
    "transaction_id": ["transaction_id", "transaction id", "id транзакции"],
    "order_id": ["order_id", "order id", "id заказа", "номер заказа", "заказ"],
    "paid_at": [
        "paid_at",
        "payment date",
        "дата платежа",
        "дата операции",
        "дата",
        "date",
    ],
    "operation_type": ["operation_type", "operation type", "тип операции", "operation"],
    "amount": ["amount", "sum", "сумма", "стоимость", "оплата"],
    "payment_method": ["payment method", "payment_method", "method", "способ оплаты"],
    "client_id": ["client_id", "client id", "customer_id", "id клиента", "клиент"],
    "product_name": ["product", "product_name", "название товара", "наименование"],
    "product_category": ["category", "product_category", "категория", "группа"],
    "manager": ["manager", "sales", "менеджер", "продавец"],
    "group_1": ["group_1", "group 1", "группа 1", "группировка 1"],
    "group_2": ["group_2", "group 2", "группа 2", "группировка 2"],
    "group_3": ["group_3", "group 3", "группа 3", "группировка 3"],
    "group_4": ["group_4", "group 4", "группа 4", "группировка 4"],
    "group_5": ["group_5", "group 5", "группа 5", "группировка 5"],
    "fee_1": ["fee_1", "fee 1", "комиссия 1"],
    "fee_2": ["fee_2", "fee 2", "комиссия 2"],
    "fee_3": ["fee_3", "fee 3", "комиссия 3"],
    "utm_source": ["utm_source", "utm source", "источник", "utm source"],
    "utm_medium": ["utm_medium", "utm medium", "utm medium"],
    "utm_campaign": ["utm_campaign", "utm campaign", "utm campaign"],
    "utm_term": ["utm_term", "utm term", "utm term"],
    "utm_content": ["utm_content", "utm content", "utm content"],
    "spend_amount": [
        "spend",
        "spend_amount",
        "cost",
        "расход",
        "затраты",
        "budget",
    ],
}

DATE_FORMATS = (
    "%Y-%m-%d",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%Y.%m.%d",
    "%d-%m-%Y",
)


def _get_upload(upload_id: int, current_user: CurrentUser, db: Session) -> Upload:
    upload = db.scalar(
        select(Upload)
        .join(Project, Upload.project_id == Project.id)
        .where(Upload.id == upload_id, Project.owner_id == current_user.id)
    )
    if not upload:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Загрузка не найдена.",
        )
    return upload


def _normalize_header(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9а-я]+", " ", value.lower())
    return re.sub(r"\s+", " ", normalized).strip()


def _serialize_cell(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value if value is not None else ""


def _infer_type(values: Iterable[object]) -> str:
    cleaned = [value for value in values if value not in (None, "")]
    if not cleaned:
        return "string"

    if all(_is_date(value) for value in cleaned):
        return "date"
    if all(_is_int(value) for value in cleaned):
        return "integer"
    if all(_is_float(value) for value in cleaned):
        return "float"
    return "string"


def _is_date(value: object) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, str):
        for fmt in DATE_FORMATS:
            try:
                datetime.strptime(value.strip(), fmt)
                return True
            except ValueError:
                continue
    return False


def _is_int(value: object) -> bool:
    if isinstance(value, int) and not isinstance(value, bool):
        return True
    if isinstance(value, str):
        return bool(re.fullmatch(r"-?\d+", value.strip()))
    return False


def _is_float(value: object) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    if isinstance(value, str):
        return bool(re.fullmatch(r"-?\d+[.,]?\d*", value.strip()))
    return False


def _read_csv_preview_with_encoding(
    file_path: Path,
    encoding: str,
    errors: str = "strict",
) -> tuple[list[str], list[list[object]]]:
    with file_path.open("r", encoding=encoding, errors=errors, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        headers = next(reader, [])
        rows: list[list[object]] = []
        for row in reader:
            rows.append([_serialize_cell(value) for value in row])
            if len(rows) >= 20:
                break
    return headers, rows


def _read_csv_preview(file_path: Path) -> tuple[list[str], list[list[object]]]:
    try:
        return _read_csv_preview_with_encoding(file_path, "utf-8-sig")
    except UnicodeDecodeError:
        return _read_csv_preview_with_encoding(
            file_path,
            "cp1251",
            errors="replace",
        )


def _read_xlsx_preview(file_path: Path) -> tuple[list[str], list[list[object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Для предпросмотра XLSX установите зависимость openpyxl.",
        ) from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    headers = [str(value) if value is not None else "" for value in (headers_row or [])]
    rows: list[list[object]] = []
    for row in rows_iter:
        rows.append([_serialize_cell(value) for value in row])
        if len(rows) >= 20:
            break
    workbook.close()
    return headers, rows


def _preview_from_upload(upload: Upload) -> tuple[list[str], list[list[object]]]:
    file_path = Path(upload.file_path)
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Файл загрузки не найден.",
        )
    if file_path.suffix.lower() == ".xlsx":
        return _read_xlsx_preview(file_path)
    return _read_csv_preview(file_path)


def _build_mapping_suggestions(
    headers: list[str],
    upload_type: UploadType,
) -> dict[str, str | None]:
    if upload_type == UploadType.TRANSACTIONS:
        available_fields = {
            "paid_at",
            "amount",
            "operation_type",
            "payment_method",
            "group_1",
            "group_2",
            "group_3",
            "group_4",
            "group_5",
            "fee_1",
            "fee_2",
            "fee_3",
            "transaction_id",
            "order_id",
            "client_id",
            "product_name",
            "product_category",
            "manager",
            "utm_source",
            "utm_medium",
            "utm_campaign",
            "utm_term",
            "utm_content",
        }
    else:
        available_fields = set(REQUIRED_FIELDS[upload_type])
    suggestions: dict[str, str | None] = {}
    for header in headers:
        normalized = _normalize_header(header)
        matched = None
        for field, keywords in SUGGESTION_RULES.items():
            if field not in available_fields:
                continue
            if any(keyword in normalized for keyword in keywords):
                matched = field
                break
        suggestions[header] = matched
    return suggestions


def _get_mapping(upload_id: int, db: Session) -> ColumnMapping:
    mapping = db.get(ColumnMapping, upload_id)
    if not mapping:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Сначала сохраните маппинг колонок.",
        )
    return mapping


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _truncate_string(value: str, max_length: int) -> str:
    if len(value) <= max_length:
        return value
    return value[:max_length]


def _normalize_operation_value(value: object) -> str:
    return _stringify(value).strip().lower()


def _infer_operation_from_payment_type(value: str) -> str | None:
    normalized = value.strip().lower()
    if not normalized:
        return None
    sale_markers = (
        "оплата",
        "приход",
        "поступление",
        "пополнение",
        "прибыль",
        "выплата",
        "оплачено",
    )
    refund_markers = ("возврат", "возвращено", "отклонено", "отмена")
    if any(marker in normalized for marker in refund_markers):
        return "refund"
    if any(marker in normalized for marker in sale_markers):
        return "sale"
    return None


def _parse_fee_value(raw_value: object) -> tuple[float, bool]:
    parsed = parse_float(raw_value)
    if parsed is None:
        if _stringify(raw_value).strip() == "":
            return 0.0, False
        return 0.0, True
    return float(parsed), False


def _build_quality_report(
    upload: Upload,
    mapping: ColumnMapping,
) -> tuple[QualityReport, list[dict[str, object]]]:
    headers, rows = read_upload_rows(upload)
    header_index = {header: index for index, header in enumerate(headers)}
    field_to_header = build_field_mapping(mapping.mapping_json)
    mapping_json = mapping.mapping_json or {}
    normalization = mapping.normalization_json or {}
    operation_type_mapping = extract_operation_type_mapping(mapping_json)
    unknown_operation_policy = extract_unknown_operation_policy(mapping_json)
    required_fields = REQUIRED_FIELDS[upload.type]

    errors: list[QualityIssue] = []
    warnings: list[QualityIssue] = []
    rows_with_errors: set[int] = set()
    skipped_rows = 0
    processed_rows: list[dict[str, object]] = []

    seen_transactions: set[str] = set()
    optional_fields = [
        "transaction_id",
        "order_id",
        "client_id",
        "product_name",
        "product_category",
        "manager",
        "payment_method",
        "group_1",
        "group_2",
        "group_3",
        "group_4",
        "group_5",
        "fee_1",
        "fee_2",
        "fee_3",
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "utm_term",
        "utm_content",
    ]

    for row_index, row in enumerate(rows, start=2):
        row_has_error = False
        row_skip = False
        row_payload: dict[str, object] = {}
        parsed_payload: dict[str, object] = {}
        row_issues: list[dict[str, str | int]] = []

        for field in required_fields:
            header = field_to_header.get(field, "")
            raw_value = get_row_value(row, header_index, header) if header else ""
            normalized_value = normalize_value(raw_value, normalization.get(header))
            row_payload[field] = {
                "raw": _stringify(raw_value),
                "normalized": normalized_value,
                "header": header,
            }
            if normalized_value in ("", None):
                errors.append(
                    QualityIssue(
                        row=row_index,
                        field=field,
                        message="Поле обязательно для заполнения.",
                    )
                )
                row_issues.append(
                    {
                        "level": "error",
                        "row": row_index,
                        "field": field,
                        "message": "Поле обязательно для заполнения.",
                    }
                )
                row_has_error = True

        if upload.type == UploadType.TRANSACTIONS:
            for field in optional_fields:
                header = field_to_header.get(field, "")
                if not header:
                    continue
                raw_value = get_row_value(row, header_index, header)
                normalized_value = normalize_value(raw_value, normalization.get(header))
                row_payload[field] = {
                    "raw": _stringify(raw_value),
                    "normalized": normalized_value,
                    "header": header,
                }

            transaction_key = _stringify(
                row_payload.get("transaction_id", {}).get("normalized", "")
            ) or _stringify(row_payload.get("order_id", {}).get("normalized", ""))
            if transaction_key:
                if transaction_key in seen_transactions:
                    warnings.append(
                        QualityIssue(
                            row=row_index,
                            field="transaction_id" if row_payload.get("transaction_id") else "order_id",
                            message="Повторяющийся transaction_id/order_id.",
                        )
                    )
                    row_issues.append(
                        {
                            "level": "warning",
                            "row": row_index,
                            "field": "transaction_id" if row_payload.get("transaction_id") else "order_id",
                            "message": "Повторяющийся transaction_id/order_id.",
                        }
                    )
                else:
                    seen_transactions.add(transaction_key)

            date_value = parse_date(
                row_payload.get("paid_at", {}).get("raw", "")
            )
            if not date_value:
                errors.append(
                    QualityIssue(
                        row=row_index,
                        field="paid_at",
                        message="Дата платежа не распознана.",
                    )
                )
                row_issues.append(
                    {
                        "level": "error",
                        "row": row_index,
                        "field": "paid_at",
                        "message": "Дата платежа не распознана.",
                    }
                )
                row_has_error = True

            amount_value = parse_float(
                row_payload.get("amount", {}).get("raw", "")
            )
            if amount_value is None or amount_value == 0:
                errors.append(
                    QualityIssue(
                        row=row_index,
                        field="amount",
                        message="Сумма должна быть больше 0.",
                    )
                )
                row_issues.append(
                    {
                        "level": "error",
                        "row": row_index,
                        "field": "amount",
                        "message": "Сумма должна быть больше 0.",
                    }
                )
                row_has_error = True
            elif amount_value < 0:
                warnings.append(
                    QualityIssue(
                        row=row_index,
                        field="amount",
                        message="Отрицательная сумма, используем модуль.",
                    )
                )
                row_issues.append(
                    {
                        "level": "warning",
                        "row": row_index,
                        "field": "amount",
                        "message": "Отрицательная сумма, используем модуль.",
                    }
                )
                amount_value = abs(amount_value)

            operation_value = _normalize_operation_value(
                row_payload.get("operation_type", {}).get("normalized", "")
            )
            resolved_operation = None
            if operation_value:
                mapped_value = operation_type_mapping.get(operation_value)
                if mapped_value:
                    resolved_operation = mapped_value
                elif operation_value in {"sale", "refund"}:
                    resolved_operation = operation_value
                else:
                    inferred_from_operation = _infer_operation_from_payment_type(
                        operation_value
                    )
                    if inferred_from_operation:
                        resolved_operation = inferred_from_operation
            if not resolved_operation:
                payment_type_value = _stringify(
                    row_payload.get("payment_method", {}).get("normalized", "")
                )
                inferred_operation = _infer_operation_from_payment_type(
                    payment_type_value
                )
                if inferred_operation:
                    resolved_operation = inferred_operation
            if not resolved_operation:
                message = "Тип операции не распознан."
                if unknown_operation_policy == "ignore":
                    warnings.append(
                        QualityIssue(
                            row=row_index,
                            field="operation_type",
                            message=f"{message} Строка пропущена.",
                        )
                    )
                    row_issues.append(
                        {
                            "level": "warning",
                            "row": row_index,
                            "field": "operation_type",
                            "message": f"{message} Строка пропущена.",
                        }
                    )
                    row_skip = True
                else:
                    errors.append(
                        QualityIssue(
                            row=row_index,
                            field="operation_type",
                            message="Тип операции должен быть sale или refund.",
                        )
                    )
                    row_issues.append(
                        {
                            "level": "error",
                            "row": row_index,
                            "field": "operation_type",
                            "message": "Тип операции должен быть sale или refund.",
                        }
                    )
                    row_has_error = True

            fee_total = 0.0
            for fee_field in ("fee_1", "fee_2", "fee_3"):
                raw_fee = row_payload.get(fee_field, {}).get("raw", "")
                fee_value, fee_invalid = _parse_fee_value(raw_fee)
                if fee_invalid:
                    warnings.append(
                        QualityIssue(
                            row=row_index,
                            field=fee_field,
                            message="Комиссия не распознана, использовано 0.",
                        )
                    )
                    row_issues.append(
                        {
                            "level": "warning",
                            "row": row_index,
                            "field": fee_field,
                            "message": "Комиссия не распознана, использовано 0.",
                        }
                    )
                parsed_payload[fee_field] = fee_value
                fee_total += fee_value

            if not row_has_error and not row_skip:
                parsed_payload["paid_at"] = date_value
                parsed_payload["amount"] = amount_value
                parsed_payload["operation_type"] = resolved_operation
                parsed_payload["fee_total"] = fee_total
        else:
            date_value = parse_date(
                row_payload.get("date", {}).get("raw", "")
            )
            if not date_value:
                errors.append(
                    QualityIssue(
                        row=row_index,
                        field="date",
                        message="Дата не распознана.",
                    )
                )
                row_issues.append(
                    {
                        "level": "error",
                        "row": row_index,
                        "field": "date",
                        "message": "Дата не распознана.",
                    }
                )
                row_has_error = True

            spend_value = parse_float(
                row_payload.get("spend_amount", {}).get("raw", "")
            )
            if spend_value is None or spend_value <= 0:
                errors.append(
                    QualityIssue(
                        row=row_index,
                        field="spend_amount",
                        message="Сумма должна быть больше 0.",
                    )
                )
                row_issues.append(
                    {
                        "level": "error",
                        "row": row_index,
                        "field": "spend_amount",
                        "message": "Сумма должна быть больше 0.",
                    }
                )
                row_has_error = True

            if not row_has_error:
                parsed_payload["date"] = date_value
                parsed_payload["spend_amount"] = spend_value

        if row_has_error:
            rows_with_errors.add(row_index)
        if row_skip and not row_has_error:
            skipped_rows += 1
        processed_rows.append(
            {
                "row_index": row_index,
                "payload": row_payload,
                "parsed": parsed_payload,
                "skip": row_has_error or row_skip,
                "issues": row_issues,
            }
        )

    report = QualityReport(
        errors=errors,
        warnings=warnings,
        stats=QualityStats(
            total_rows=len(rows),
            valid_rows=len(rows) - len(rows_with_errors) - skipped_rows,
            error_count=len(errors),
            warning_count=len(warnings),
            skipped_rows=skipped_rows,
        ),
    )
    return report, processed_rows


@router.get("/{upload_id}/preview", response_model=UploadPreview)
def preview_upload(
    upload_id: int,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
) -> UploadPreview:
    upload = _get_upload(upload_id, current_user, db)
    headers, sample_rows = _preview_from_upload(upload)
    inferred_types = {
        header: _infer_type(row[index] if index < len(row) else "" for row in sample_rows)
        for index, header in enumerate(headers)
    }
    mapping_suggestions = _build_mapping_suggestions(headers, upload.type)
    column_stats: dict[str, dict[str, object]] = {}
    for index, header in enumerate(headers):
        values = [
            _stringify(row[index]) if index < len(row) else ""
            for row in sample_rows
        ]
        unique_values: list[str] = []
        seen: set[str] = set()
        for value in values:
            if value in seen:
                continue
            seen.add(value)
            unique_values.append(value)
        column_stats[header] = {
            "unique_values": unique_values,
            "unique_count": len(unique_values),
            "sample_count": len(values),
        }
    return UploadPreview(
        headers=headers,
        sample_rows=sample_rows,
        inferred_types=inferred_types,
        mapping_suggestions=mapping_suggestions,
        column_stats=column_stats,
        upload_type=upload.type,
        project_id=upload.project_id,
    )


@router.post(
    "/{upload_id}/mapping",
    response_model=ColumnMappingPublic,
    status_code=status.HTTP_201_CREATED,
)
def save_mapping(
    upload_id: int,
    payload: ColumnMappingCreate,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
) -> ColumnMappingPublic:
    upload = _get_upload(upload_id, current_user, db)
    required_fields = REQUIRED_FIELDS[upload.type]
    selected_fields = {value for value in payload.mapping.values() if value}
    missing = [field for field in required_fields if field not in selected_fields]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Не заполнены обязательные поля: {', '.join(missing)}.",
        )
    if payload.unknown_operation_policy not in {"error", "ignore"}:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Неверная политика для нераспознанных типов операций.",
        )

    mapping = db.get(ColumnMapping, upload_id)
    normalization = payload.normalization or {}
    operation_mapping = {
        _normalize_operation_value(key): value
        for key, value in (payload.operation_type_mapping or {}).items()
        if _normalize_operation_value(key)
    }
    invalid_operation_values = {
        value for value in operation_mapping.values() if value not in {"sale", "refund"}
    }
    if invalid_operation_values:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Неверные значения для маппинга типов операций.",
        )
    mapping_payload = {
        "mapping": payload.mapping,
        "value_mapping": {"operation_type": operation_mapping},
        "unknown_operation_policy": payload.unknown_operation_policy,
    }
    if mapping:
        mapping.mapping_json = mapping_payload
        mapping.normalization_json = normalization
        mapping.created_at = datetime.now(timezone.utc)
    else:
        mapping = ColumnMapping(
            upload_id=upload_id,
            mapping_json=mapping_payload,
            normalization_json=normalization,
        )
        db.add(mapping)
    db.commit()
    db.refresh(mapping)
    return ColumnMappingPublic.model_validate(mapping)


@router.post("/{upload_id}/validate", response_model=QualityReport)
def validate_upload(
    upload_id: int,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
) -> QualityReport:
    upload = _get_upload(upload_id, current_user, db)
    mapping = _get_mapping(upload_id, db)
    report, _ = _build_quality_report(upload, mapping)
    upload.status = UploadStatus.VALIDATED if not report.errors else UploadStatus.FAILED
    db.commit()
    return report


@router.post("/{upload_id}/import", response_model=ImportResult)
def import_upload(
    upload_id: int,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
) -> ImportResult:
    upload = _get_upload(upload_id, current_user, db)
    mapping = _get_mapping(upload_id, db)
    report, processed_rows = _build_quality_report(upload, mapping)
    inserted = 0
    normalization = mapping.normalization_json or {}
    mapping_json = mapping.mapping_json or {}
    field_to_header = build_field_mapping(mapping_json)
    settings = db.get(ProjectSettings, upload.project_id)
    dedup_policy = settings.dedup_policy if settings else "keep_all_rows"

    quarantine_rows: list[UploadQuarantineRow] = []
    ready_rows: list[dict[str, object]] = []
    for row_entry in processed_rows:
        if row_entry.get("skip"):
            quarantine_rows.append(
                UploadQuarantineRow(
                    upload_id=upload.id,
                    row_number=row_entry.get("row_index", 0),
                    issues_json=row_entry.get("issues", []),
                    payload_json=row_entry.get("payload", {}),
                )
            )
            continue
        ready_rows.append(row_entry)

    def dedup_key(entry: dict[str, object], use_order_id: bool = True) -> str | None:
        payload = entry.get("payload", {})
        if not isinstance(payload, dict):
            return None
        transaction_id = _stringify(
            payload.get("transaction_id", {}).get("normalized", "")
        )
        if transaction_id:
            return transaction_id
        if use_order_id:
            order_id = _stringify(payload.get("order_id", {}).get("normalized", ""))
            return order_id or None
        return None

    if dedup_policy == "last_row_wins":
        deduped: dict[str, dict[str, object]] = {}
        passthrough: list[dict[str, object]] = []
        for row_entry in ready_rows:
            key = dedup_key(row_entry)
            if not key:
                passthrough.append(row_entry)
                continue
            deduped[key] = row_entry
        ready_rows = passthrough + list(deduped.values())
    elif dedup_policy == "aggregate_by_transaction_id":
        aggregated: dict[str, dict[str, object]] = {}
        passthrough: list[dict[str, object]] = []
        for row_entry in ready_rows:
            key = dedup_key(row_entry, use_order_id=False)
            if not key:
                passthrough.append(row_entry)
                continue
            if key not in aggregated:
                aggregated[key] = row_entry
                continue
            base = aggregated[key]
            base_parsed = base.get("parsed", {})
            row_parsed = row_entry.get("parsed", {})
            if isinstance(base_parsed, dict) and isinstance(row_parsed, dict):
                base_parsed["amount"] = float(base_parsed.get("amount", 0.0)) + float(
                    row_parsed.get("amount", 0.0)
                )
                base_parsed["fee_1"] = float(base_parsed.get("fee_1", 0.0)) + float(
                    row_parsed.get("fee_1", 0.0)
                )
                base_parsed["fee_2"] = float(base_parsed.get("fee_2", 0.0)) + float(
                    row_parsed.get("fee_2", 0.0)
                )
                base_parsed["fee_3"] = float(base_parsed.get("fee_3", 0.0)) + float(
                    row_parsed.get("fee_3", 0.0)
                )
                base_parsed["fee_total"] = float(base_parsed.get("fee_total", 0.0)) + float(
                    row_parsed.get("fee_total", 0.0)
                )
        ready_rows = passthrough + list(aggregated.values())

    for row_entry in ready_rows:
        row_payload = row_entry.get("payload", {})
        parsed_payload = row_entry.get("parsed", {})
        if upload.type == UploadType.TRANSACTIONS:
            product_header = field_to_header.get("product_name", "")
            manager_header = field_to_header.get("manager", "")
            product_raw = row_payload.get("product_name", {}).get("raw", "")
            manager_raw = row_payload.get("manager", {}).get("raw", "")
            product_key = normalize_value(
                product_raw, normalization.get(product_header)
            )
            manager_key = normalize_value(
                manager_raw, normalization.get(manager_header)
            )
            product_id = (
                resolve_product_alias(
                    db, upload.project_id, _stringify(product_key)
                )
                if _stringify(product_key)
                else None
            )
            manager_id = (
                resolve_manager_alias(
                    db, upload.project_id, _stringify(manager_key)
                )
                if _stringify(manager_key)
                else None
            )
            product_norm = (
                get_product_name(db, product_id)
                if product_id
                else _stringify(product_key) or None
            )
            manager_norm = (
                get_manager_name(db, manager_id)
                if manager_id
                else _stringify(manager_key) or None
            )
            record = FactTransaction(
                project_id=upload.project_id,
                transaction_id=_truncate_string(
                    _stringify(
                        row_payload.get("transaction_id", {}).get("normalized", "")
                    ),
                    128,
                )
                or None,
                order_id=_truncate_string(
                    _stringify(row_payload.get("order_id", {}).get("normalized", "")),
                    128,
                )
                or None,
                date=parsed_payload.get("paid_at"),
                operation_type=_truncate_string(
                    _stringify(parsed_payload.get("operation_type")), 32
                ),
                amount=parsed_payload.get("amount"),
                client_id=_truncate_string(
                    _stringify(row_payload.get("client_id", {}).get("normalized", "")),
                    128,
                )
                or None,
                product_name_raw=_truncate_string(_stringify(product_raw), 255)
                or None,
                product_name_norm=_truncate_string(_stringify(product_norm), 255)
                or None,
                product_id=product_id,
                product_category=_truncate_string(
                    _stringify(
                        row_payload.get("product_category", {}).get("normalized", "")
                    ),
                    255,
                )
                or None,
                manager_raw=_truncate_string(_stringify(manager_raw), 255) or None,
                manager_norm=_truncate_string(_stringify(manager_norm), 255) or None,
                manager_id=manager_id,
                payment_method=_truncate_string(
                    _stringify(
                        row_payload.get("payment_method", {}).get("normalized", "")
                    ),
                    255,
                )
                or None,
                group_1=_truncate_string(
                    _stringify(row_payload.get("group_1", {}).get("normalized", "")),
                    255,
                )
                or None,
                group_2=_truncate_string(
                    _stringify(row_payload.get("group_2", {}).get("normalized", "")),
                    255,
                )
                or None,
                group_3=_truncate_string(
                    _stringify(row_payload.get("group_3", {}).get("normalized", "")),
                    255,
                )
                or None,
                group_4=_truncate_string(
                    _stringify(row_payload.get("group_4", {}).get("normalized", "")),
                    255,
                )
                or None,
                group_5=_truncate_string(
                    _stringify(row_payload.get("group_5", {}).get("normalized", "")),
                    255,
                )
                or None,
                fee_1=parsed_payload.get("fee_1"),
                fee_2=parsed_payload.get("fee_2"),
                fee_3=parsed_payload.get("fee_3"),
                fee_total=parsed_payload.get("fee_total"),
                commission=parsed_payload.get("fee_total"),
                utm_source=_truncate_string(
                    _stringify(row_payload.get("utm_source", {}).get("normalized", "")),
                    255,
                )
                or None,
                utm_medium=_truncate_string(
                    _stringify(row_payload.get("utm_medium", {}).get("normalized", "")),
                    255,
                )
                or None,
                utm_campaign=_truncate_string(
                    _stringify(
                        row_payload.get("utm_campaign", {}).get("normalized", "")
                    ),
                    255,
                )
                or None,
                utm_term=_truncate_string(
                    _stringify(row_payload.get("utm_term", {}).get("normalized", "")),
                    255,
                )
                or None,
                utm_content=_truncate_string(
                    _stringify(
                        row_payload.get("utm_content", {}).get("normalized", "")
                    ),
                    255,
                )
                or None,
            )
            db.add(record)
        else:
            record = FactMarketingSpend(
                project_id=upload.project_id,
                date=parsed_payload.get("date"),
                spend_amount=parsed_payload.get("spend_amount"),
            )
            db.add(record)
        inserted += 1

    upload.status = UploadStatus.IMPORTED
    if quarantine_rows:
        db.add_all(quarantine_rows)
    db.commit()
    try:
        generate_insights_for_project(db, upload.project_id)
        db.commit()
    except Exception:
        db.rollback()
    return ImportResult(imported=inserted)
