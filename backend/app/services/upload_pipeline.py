from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status

from app.models.upload import Upload

DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%Y.%m.%d",
    "%d-%m-%Y",
)


def read_upload_rows(upload: Upload) -> tuple[list[str], list[list[Any]]]:
    file_path = Path(upload.file_path)
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Файл загрузки не найден.",
        )
    if file_path.suffix.lower() == ".xlsx":
        return _read_xlsx_rows(file_path)
    return _read_csv_rows(file_path)


def _read_csv_rows_with_encoding(
    file_path: Path,
    encoding: str,
    errors: str = "strict",
) -> tuple[list[str], list[list[Any]]]:
    with file_path.open("r", encoding=encoding, errors=errors, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        headers = next(reader, [])
        rows = [list(row) for row in reader]
    return headers, rows


def _read_csv_rows(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        return _read_csv_rows_with_encoding(file_path, "utf-8-sig")
    except UnicodeDecodeError:
        return _read_csv_rows_with_encoding(file_path, "cp1251", errors="replace")


def _read_xlsx_rows(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Для обработки XLSX установите зависимость openpyxl.",
        ) from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    headers = [str(value) if value is not None else "" for value in (headers_row or [])]
    rows: list[list[Any]] = []
    for row in rows_iter:
        rows.append([_serialize_cell(value) for value in row])
    workbook.close()
    return headers, rows


def _serialize_cell(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value if value is not None else ""


def normalize_value(value: Any, rules: dict[str, Any] | None) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    normalized = value
    if rules:
        if rules.get("trim"):
            normalized = normalized.strip()
        if rules.get("lowercase"):
            normalized = normalized.lower()
        if rules.get("uppercase"):
            normalized = normalized.upper()
    return normalized


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        cleaned = value.strip()
        try:
            return datetime.fromisoformat(cleaned).date()
        except ValueError:
            pass
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\u00a0", " ").strip()
        cleaned = cleaned.replace(" ", "")
        cleaned = re.sub(r"[^\d,.\-]", "", cleaned)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def extract_mapping(mapping_json: dict[str, Any]) -> dict[str, str | None]:
    if "mapping" in mapping_json and isinstance(mapping_json["mapping"], dict):
        return mapping_json["mapping"]
    return mapping_json  # type: ignore[return-value]


def extract_operation_type_mapping(mapping_json: dict[str, Any]) -> dict[str, str]:
    value_mapping = mapping_json.get("value_mapping", {})
    if not isinstance(value_mapping, dict):
        return {}
    operation_mapping = value_mapping.get("operation_type", {})
    if isinstance(operation_mapping, dict):
        return {str(key): str(value) for key, value in operation_mapping.items()}
    return {}


def extract_unknown_operation_policy(mapping_json: dict[str, Any]) -> str:
    policy = mapping_json.get("unknown_operation_policy", "error")
    if policy not in {"error", "ignore"}:
        return "error"
    return policy


def build_field_mapping(mapping_json: dict[str, Any]) -> dict[str, str]:
    field_to_header: dict[str, str] = {}
    for header, field in extract_mapping(mapping_json).items():
        if not field or field in {"ignore", "not_set"}:
            continue
        if field not in field_to_header:
            field_to_header[field] = header
    return field_to_header


def get_row_value(
    row: list[Any],
    header_index: dict[str, int],
    header: str,
) -> Any:
    index = header_index.get(header)
    if index is None or index >= len(row):
        return ""
    return row[index]
